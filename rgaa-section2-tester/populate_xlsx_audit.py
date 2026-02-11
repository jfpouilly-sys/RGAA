#!/usr/bin/env python3
"""
RGAA CSV to XLSX Audit Grid Populator
Updates RGAA audit XLSX template with results from CSV files.
Uses openpyxl for fast, reliable manipulation of Excel files.
"""

import csv
from pathlib import Path
import argparse
import shutil
from typing import Dict, List, Optional
import sys

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Error: openpyxl not installed. Install with: pip install openpyxl --break-system-packages")


class XLSXAuditPopulator:
    """Populates XLSX audit grid with data from CSV files using openpyxl"""
    
    # Status mapping from CSV to XLSX
    STATUS_MAPPING = {
        'C': 'C',      # Conforme
        'NC': 'NC',    # Non Conforme
        'NA': 'NA',    # Non Applicable
        'NT': 'NT'     # Non Testé
    }
    
    def __init__(self, xlsx_template: Path, output_xlsx: Optional[Path] = None):
        self.xlsx_template = xlsx_template
        self.output_xlsx = output_xlsx or xlsx_template.parent / f"{xlsx_template.stem}_updated.xlsx"
        self.csv_data: Dict[str, List[Dict]] = {}
        
    def load_csv_files(self, csv_pattern: str) -> int:
        """Load all CSV files matching the pattern"""
        csv_files = sorted(Path('.').glob(csv_pattern))
        
        if not csv_files:
            print(f"Warning: No CSV files found matching pattern: {csv_pattern}")
            return 0
        
        for csv_file in csv_files:
            page_id = csv_file.stem
            
            try:
                data = self._read_csv_file(csv_file)
                self.csv_data[page_id] = data
                print(f"  ✓ Loaded {csv_file.name} ({len(data)} criteria)")
            except Exception as e:
                print(f"  ✗ Error loading {csv_file.name}: {e}")
        
        return len(self.csv_data)
    
    def load_single_csv(self, csv_file: Path) -> bool:
        """Load a single CSV file"""
        page_id = csv_file.stem
        
        try:
            data = self._read_csv_file(csv_file)
            self.csv_data[page_id] = data
            print(f"  ✓ Loaded {csv_file.name} ({len(data)} criteria)")
            return True
        except Exception as e:
            print(f"  ✗ Error loading {csv_file.name}: {e}")
            return False
    
    def _read_csv_file(self, csv_file: Path) -> List[Dict]:
        """Read CSV file and return list of dictionaries"""
        data = []
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                data.append(row)
        return data
    
    def update_xlsx_file(self) -> bool:
        """Update the XLSX file with data from loaded CSV files"""
        if not HAS_OPENPYXL:
            print("Error: openpyxl is required but not installed")
            return False
            
        if not self.csv_data:
            print("Error: No CSV data loaded")
            return False
        
        # Copy template to output location
        try:
            shutil.copy2(self.xlsx_template, self.output_xlsx)
            print(f"\n✓ Created working copy: {self.output_xlsx.name}")
        except Exception as e:
            print(f"Error copying template: {e}")
            return False
        
        # Load XLSX workbook with data_only=True to get calculated formula values
        try:
            print("Loading XLSX workbook...")
            wb = load_workbook(str(self.output_xlsx), data_only=True)
        except Exception as e:
            print(f"Error loading XLSX: {e}")
            return False
        
        # Process each page
        success_count = 0
        for page_id, csv_rows in self.csv_data.items():
            print(f"Processing {page_id}...", end=' ')
            if self._update_worksheet(wb, page_id, csv_rows):
                success_count += 1
        
        # Save updated workbook
        try:
            print("\nSaving updated XLSX file...")
            wb.save(str(self.output_xlsx))
            print(f"✓ Saved successfully")
        except Exception as e:
            print(f"Error saving XLSX file: {e}")
            return False
        
        print(f"\n{'='*60}")
        print(f"Update complete: {success_count}/{len(self.csv_data)} pages updated")
        print(f"Output file: {self.output_xlsx}")
        
        return success_count > 0
    
    def _update_worksheet(self, wb, page_id: str, csv_rows: List[Dict]) -> bool:
        """Update a single worksheet in the XLSX workbook"""
        try:
            # Check if sheet exists
            if page_id not in wb.sheetnames:
                print(f"\n  ⚠ Sheet '{page_id}' not found in XLSX template")
                return False
            
            ws = wb[page_id]
            
            # Create status mapping from CSV
            csv_status_map = {}
            page_name = ""
            page_url = ""
            
            for row in csv_rows:
                criterion = str(row['Critère']).strip()
                status = str(row['Statut']).strip()
                csv_status_map[criterion] = self.STATUS_MAPPING.get(status, status)
                
                # Get page info from first row
                if not page_name:
                    page_name = row['Page']
                    page_url = row['URL']
            
            # Update row 2 (page info) - Excel is 1-indexed
            if page_name and page_url:
                ws.cell(row=2, column=1).value = f"{page_name} : {page_url}"
            
            # Update data rows (starting from row 5 in Excel - row index 5)
            # Column mapping: A=1(Thématique/empty), B=2(Critère), C=3(Recommandation), 
            #                 D=4(Statut), E=5(Dérogation)
            updates = 0
            
            # Iterate through rows starting from row 5
            for row_idx in range(5, ws.max_row + 1):
                # Column B (2) contains the criterion number in data rows
                criterion_cell = ws.cell(row=row_idx, column=2)
                criterion_text = str(criterion_cell.value).strip() if criterion_cell.value else ""
                
                if not criterion_text or criterion_text not in csv_status_map:
                    continue
                
                # Update Status (column D = 4)
                new_status = csv_status_map[criterion_text]
                ws.cell(row=row_idx, column=4).value = new_status
                
                # Update Dérogation (column E = 5)
                if new_status == 'NA':
                    ws.cell(row=row_idx, column=5).value = 'S.O.'
                elif new_status in ['C', 'NC', 'NT']:
                    ws.cell(row=row_idx, column=5).value = 'N'
                
                updates += 1
            
            print(f"✓ {updates} criteria updated")
            return True
            
        except Exception as e:
            print(f"\n  ✗ Error updating {page_id}: {e}")
            import traceback
            traceback.print_exc()
            return False


def populate_from_csv(csv_file: Path, xlsx_template: Path, 
                     output_xlsx: Optional[Path] = None) -> bool:
    """Populate XLSX template with data from a single CSV file"""
    populator = XLSXAuditPopulator(xlsx_template, output_xlsx)
    
    if not populator.load_single_csv(csv_file):
        return False
    
    return populator.update_xlsx_file()


def batch_populate(csv_pattern: str, xlsx_template: Path,
                   output_xlsx: Optional[Path] = None) -> bool:
    """Populate XLSX template with data from multiple CSV files"""
    populator = XLSXAuditPopulator(xlsx_template, output_xlsx)
    
    num_loaded = populator.load_csv_files(csv_pattern)
    if num_loaded == 0:
        return False
    
    return populator.update_xlsx_file()


def main():
    """Main entry point for command-line usage"""
    parser = argparse.ArgumentParser(
        description='Populate RGAA XLSX audit grid with data from CSV files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Update template with a single CSV file
  python populate_xlsx_audit.py template.xlsx P01.csv
  
  # Update template with all P*.csv files
  python populate_xlsx_audit.py template.xlsx "P*.csv"
  
  # Specify custom output filename
  python populate_xlsx_audit.py template.xlsx "P*.csv" -o audit_2026.xlsx
        """
    )
    
    parser.add_argument(
        'xlsx_template',
        type=Path,
        help='Path to XLSX template file'
    )
    parser.add_argument(
        'csv_input',
        help='CSV file or glob pattern (e.g., P01.csv or "P*.csv")'
    )
    parser.add_argument(
        '-o', '--output',
        type=Path,
        help='Output XLSX file (default: template_updated.xlsx)'
    )
    
    args = parser.parse_args()
    
    # Check if template exists
    if not args.xlsx_template.exists():
        print(f"Error: XLSX template not found: {args.xlsx_template}")
        return 1
    
    print(f"RGAA XLSX Audit Grid Populator")
    print("=" * 60)
    print(f"Template: {args.xlsx_template}")
    print(f"CSV Input: {args.csv_input}")
    if args.output:
        print(f"Output: {args.output}")
    print("-" * 60)
    
    # Check if input is a pattern or single file
    input_path = Path(args.csv_input)
    
    if '*' in args.csv_input or '?' in args.csv_input:
        # Batch mode
        print(f"\nLoading CSV files matching: {args.csv_input}")
        success = batch_populate(args.csv_input, args.xlsx_template, args.output)
    else:
        # Single file mode
        if not input_path.exists():
            print(f"Error: CSV file not found: {args.csv_input}")
            return 1
        
        print(f"\nLoading CSV file: {input_path.name}")
        success = populate_from_csv(input_path, args.xlsx_template, args.output)
    
    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
